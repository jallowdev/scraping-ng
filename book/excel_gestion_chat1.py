# StockDashboardGenerator.py
import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.worksheet.table import Table, TableStyleInfo

# Données simulées pour l'exemple
def generer_donnees():
    produits = ['Chaussures', 'T-shirt', 'Jean', 'Casquette', 'Veste']
    categories = ['Vêtements', 'Accessoires']
    ventes = []
    stock = []

    for i in range(1, 101):
        produit = np.random.choice(produits)
        categorie = 'Accessoires' if produit == 'Casquette' else 'Vêtements'
        quantite_stock = np.random.randint(5, 100)
        quantite_vendue = np.random.randint(1, 10)
        date_vente = datetime(2024, np.random.randint(1, 13), np.random.randint(1, 28))

        ventes.append({
            'Produit': produit,
            'Catégorie': categorie,
            'Quantité vendue': quantite_vendue,
            'Date de vente': date_vente
        })
        stock.append({
            'Produit': produit,
            'Catégorie': categorie,
            'Quantité en stock': quantite_stock
        })

    return pd.DataFrame(stock), pd.DataFrame(ventes)

# Création du fichier Excel avec dashboard
def creer_dashboard_excel():
    stock_df, ventes_df = generer_donnees()

    wb = Workbook()
    ws_stock = wb.active
    ws_stock.title = 'Stock'

    for r in dataframe_to_rows(stock_df, index=False, header=True):
        ws_stock.append(r)

    tab = Table(displayName="Table_Stock", ref=f"A1:C{len(stock_df)+1}")
    style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    tab.tableStyleInfo = style
    ws_stock.add_table(tab)

    ws_ventes = wb.create_sheet('Ventes')
    for r in dataframe_to_rows(ventes_df, index=False, header=True):
        ws_ventes.append(r)

    tab = Table(displayName="Table_Ventes", ref=f"A1:D{len(ventes_df)+1}")
    style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    tab.tableStyleInfo = style
    ws_ventes.add_table(tab)

    ws_dashboard = wb.create_sheet('Dashboard')
    ws_dashboard["A1"] = "Dashboard de Gestion de Stock"
    ws_dashboard["A1"].font = Font(size=14, bold=True)

    # Statistiques par produit
    ventes_par_produit = ventes_df.groupby('Produit')['Quantité vendue'].sum()
    stock_par_produit = stock_df.groupby('Produit')['Quantité en stock'].sum()

    ligne = 3
    ws_dashboard["A{}".format(ligne)] = "Ventes par produit"
    ligne += 1
    for i, (prod, qte) in enumerate(ventes_par_produit.items(), start=ligne):
        ws_dashboard[f"A{i}"] = prod
        ws_dashboard[f"B{i}"] = qte

    chart = BarChart()
    chart.title = "Ventes par Produit"
    chart.add_data(Reference(ws_dashboard, min_col=2, min_row=ligne, max_row=ligne + len(ventes_par_produit) - 1), titles_from_data=False)
    chart.set_categories(Reference(ws_dashboard, min_col=1, min_row=ligne, max_row=ligne + len(ventes_par_produit) - 1))
    ws_dashboard.add_chart(chart, "D3")

    ligne += len(ventes_par_produit) + 2
    ws_dashboard["A{}".format(ligne)] = "Stock par produit"
    ligne += 1
    for i, (prod, qte) in enumerate(stock_par_produit.items(), start=ligne):
        ws_dashboard[f"A{i}"] = prod
        ws_dashboard[f"B{i}"] = qte

    chart2 = BarChart()
    chart2.title = "Stock par Produit"
    chart2.add_data(Reference(ws_dashboard, min_col=2, min_row=ligne, max_row=ligne + len(stock_par_produit) - 1), titles_from_data=False)
    chart2.set_categories(Reference(ws_dashboard, min_col=1, min_row=ligne, max_row=ligne + len(stock_par_produit) - 1))
    ws_dashboard.add_chart(chart2, "D20")

    # Sauvegarder le fichier
    wb.save("dashboard_stock_boutique.xlsx")

if __name__ == '__main__':
    creer_dashboard_excel()
